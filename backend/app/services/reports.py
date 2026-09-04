from io import BytesIO
from xml.sax.saxutils import escape
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _paragraph(text, style):
    return Paragraph(escape(str(text)).replace("\n", "<br/>"), style)


def pdf_report(title, data):
    buf=BytesIO()
    doc=SimpleDocTemplate(
        buf,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=42,
        bottomMargin=42
    )

    styles=getSampleStyleSheet()
    title_style=ParagraphStyle(
        "ReportTitle", parent=styles["Title"],
        fontSize=20, leading=24, spaceAfter=8, alignment=TA_CENTER
    )
    sub=ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#5f6b7a"), spaceAfter=16
    )
    h=ParagraphStyle(
        "Section", parent=styles["Heading2"],
        fontSize=13, leading=16, spaceBefore=10, spaceAfter=7,
        textColor=colors.HexColor("#20304a")
    )
    body=ParagraphStyle(
        "Body", parent=styles["BodyText"],
        fontSize=9.5, leading=14, spaceAfter=5
    )
    small=ParagraphStyle(
        "Small", parent=body,
        fontSize=8.5, leading=12
    )

    story=[
        _paragraph(title,title_style),
        _paragraph(
            f"Generated for {data.get('generated_for','Learner')} ? "
            "Evidence-based report from stored platform records",
            sub
        )
    ]

    def add_table(items):
        rows=[
            [_paragraph("Metric",small),_paragraph("Result",small)]
        ]
        for item in items:
            rows.append([
                _paragraph(item.get("label",""),small),
                _paragraph(item.get("value",""),small)
            ])

        table=Table(
            rows,
            colWidths=[170,330],
            repeatRows=1
        )
        table.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#e9eef6")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#1d2a3d")),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#ccd4df")),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("LEFTPADDING",(0,0),(-1,-1),7),
            ("RIGHTPADDING",(0,0),(-1,-1),7),
            ("TOPPADDING",(0,0),(-1,-1),6),
            ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ]))
        story.append(table)
        story.append(Spacer(1,7))

    # Learner-style reports already provide explicit sections.
    sections=data.get("sections",[])
    if sections:
        for sec in sections:
            story.append(_paragraph(sec.get("title","Section"),h))

            if sec.get("items"):
                add_table(sec["items"])

            if sec.get("text"):
                story.append(_paragraph(sec["text"],body))
                story.append(Spacer(1,5))

    # Coach reports provide summary + learners.
    elif isinstance(data.get("summary"),dict):
        summary=data["summary"]

        story.append(_paragraph("Summary",h))

        summary_items=[
            {"label":"Learners","value":str(summary.get("learners",0))},
            {"label":"Assessed","value":str(summary.get("assessed",0))},
            {
                "label":"Average score",
                "value":(
                    f"{float(summary['average_score']):.1f}/100"
                    if summary.get("average_score") is not None
                    else "No assessment"
                )
            },
            {"label":"Needs coaching","value":str(summary.get("needs_coaching",0))}
        ]

        for skill,value in summary.get("skill_averages",{}).items():
            summary_items.append({
                "label":skill,
                "value":f"{float(value):.1f}/100"
            })

        add_table(summary_items)

        learners=data.get("learners",[])
        if learners:
            story.append(_paragraph("Learners",h))

            rows=[
                [
                    _paragraph("Learner",small),
                    _paragraph("Email",small),
                    _paragraph("Latest score",small),
                    _paragraph("Gaps",small),
                    _paragraph("Assessments",small)
                ]
            ]

            for learner in learners:
                gaps=learner.get("gaps") or []
                gap_text=", ".join(
                    f"{g.get('skill','Unknown')} ({float(g.get('score',0)):.1f})"
                    for g in gaps
                ) or "None"

                rows.append([
                    _paragraph(learner.get("name",""),small),
                    _paragraph(learner.get("email",""),small),
                    _paragraph(
                        f"{float(learner['latest_score']):.1f}/100"
                        if learner.get("latest_score") is not None
                        else "No assessment",
                        small
                    ),
                    _paragraph(gap_text,small),
                    _paragraph(str(learner.get("scores_count",0)),small)
                ])

            table=Table(
                rows,
                colWidths=[90,145,75,130,60],
                repeatRows=1
            )
            table.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#e9eef6")),
                ("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#1d2a3d")),
                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#ccd4df")),
                ("VALIGN",(0,0),(-1,-1),"TOP"),
                ("LEFTPADDING",(0,0),(-1,-1),5),
                ("RIGHTPADDING",(0,0),(-1,-1),5),
                ("TOPPADDING",(0,0),(-1,-1),6),
                ("BOTTOMPADDING",(0,0),(-1,-1),6),
            ]))
            story.append(table)
            story.append(Spacer(1,7))

    doc.build(story)
    buf.seek(0)
    return buf


def excel_report(title,data):
    wb=Workbook();ws=wb.active;ws.title='Report'
    ws.append([title]);ws['A1'].font=Font(size=18,bold=True)
    ws.append([f"Generated for {data.get('generated_for','Learner')}"])
    ws.append([])
    for sec in data.get('sections',[]):
        ws.append([sec.get('title','Section')]);ws.cell(ws.max_row,1).font=Font(bold=True,size=13)
        for item in sec.get('items',[]): ws.append([item.get('label',''),item.get('value','')])
        if sec.get('text'): ws.append(['Details',sec['text']])
        ws.append([])
    ws.column_dimensions['A'].width=32;ws.column_dimensions['B'].width=95
    for row in ws.iter_rows():
        for c in row:
            c.alignment=Alignment(vertical='top',wrap_text=True)
    buf=BytesIO();wb.save(buf);buf.seek(0);return buf
