"""
Reports & Export (spec section 13).

Builds a combined progress report - performance score, component
breakdown, latest coaching summary, and recent activity - as either a
PDF or an Excel workbook. Also builds a per-item PDF for a single past
analysis or session (Segment 17).

Uses fpdf2 (pure Python, no system-level dependencies like the
Pango/Cairo that weasyprint would need) and openpyxl (the standard
pure-Python xlsx library). Both were checked for a clean pip install
- zero compiled dependencies in either's tree - before being added to
this project, same discipline as every other dependency here.

PDF visual design deliberately mirrors the app's own theme (indigo
accent, the same score-bar metaphor used on the Performance Score
page) rather than being plain black-on-white text, using fpdf2's
fill/rect/line drawing primitives - no extra dependency needed for
that.
"""
from datetime import datetime, timezone
from io import BytesIO
from typing import Optional

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Font

from app.core import supabase_client
from app.services.dashboard_service import get_recent_activity
from app.services.performance_scoring_service import compute_performance_score

TOOL_SHEETS = [
    ("Argument Analyses", "argument_analyses", ["created_at", "topic", "overall_score", "summary_feedback"]),
    ("Fallacy Checks", "fallacy_detections", ["created_at", "topic", "credibility_score", "reasoning_analysis"]),
    ("Counterarguments", "counterarguments", ["created_at", "topic"]),
    ("Case Reviews", "case_reviews", ["created_at", "topic", "synthesis"]),
    ("Presentations", "presentation_analyses", ["created_at", "topic", "overall_score", "summary_feedback"]),
    ("Debates", "debate_sessions", ["created_at", "topic", "user_position", "ai_position", "status", "round_count"]),
]

# Matches the app's own light-mode theme tokens (index.css) so a
# downloaded report looks like it belongs to the same product, not a
# generic default-font PDF.
ACCENT_RGB = (91, 79, 224)
ACCENT_SOFT_RGB = (236, 233, 253)
INK_RGB = (20, 23, 31)
FAINT_RGB = (91, 100, 114)
LINE_RGB = (221, 225, 235)
OK_RGB = (47, 143, 85)


class ClashLabPDF(FPDF):
    """Adds a consistent footer to every page - fpdf2's idiomatic way
    to guarantee it appears even if content overflows onto page 2+."""

    def footer(self):
        self.set_y(-15)
        self.set_draw_color(*LINE_RGB)
        self.set_line_width(0.2)
        self.line(12, self.get_y(), self.w - 12, self.get_y())
        self.set_font("Helvetica", "", 7)
        self.set_text_color(*FAINT_RGB)
        self.set_y(-12)
        self.cell(0, 10, f"ClashLab  -  Page {self.page_no()}", align="C")
        self.set_text_color(*INK_RGB)


def _draw_header(pdf: FPDF, title: str, subtitle: str) -> None:
    pdf.set_fill_color(*ACCENT_RGB)
    pdf.rect(0, 0, pdf.w, 30, style="F")
    pdf.set_xy(12, 7)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 5, "CLASHLAB", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_xy(12, 13)
    pdf.set_font("Helvetica", "B", 17)
    pdf.cell(0, 9, _pdf_safe(title), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(*INK_RGB)
    pdf.set_xy(12, 35)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*FAINT_RGB)
    pdf.cell(0, 5, _pdf_safe(subtitle), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(*INK_RGB)
    pdf.set_y(44)


def _section_heading(pdf: FPDF, text: str) -> None:
    pdf.ln(2)
    y = pdf.get_y()
    pdf.set_fill_color(*ACCENT_RGB)
    pdf.rect(12, y + 1, 2.2, 5, style="F")
    pdf.set_xy(17, y)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(*INK_RGB)
    pdf.cell(0, 7, _pdf_safe(text), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(1)


def _score_badge(pdf: FPDF, label: str, value_text: str) -> None:
    y = pdf.get_y()
    box_w = pdf.w - 24
    pdf.set_fill_color(*ACCENT_SOFT_RGB)
    pdf.rect(12, y, box_w, 22, style="F")
    pdf.set_xy(12, y + 3)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*FAINT_RGB)
    pdf.cell(box_w, 5, _pdf_safe(label.upper()), align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_xy(12, y + 9)
    pdf.set_font("Helvetica", "B", 22)
    pdf.set_text_color(*ACCENT_RGB)
    pdf.cell(box_w, 10, value_text, align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(*INK_RGB)
    pdf.set_y(y + 26)


def _score_bar_row(pdf: FPDF, label: str, score, has_data: bool) -> None:
    y = pdf.get_y()
    label_w = 78
    bar_x = 12 + label_w
    bar_w = pdf.w - 24 - label_w - 22
    bar_h = 3.2

    pdf.set_xy(12, y + 1)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*INK_RGB)
    pdf.cell(label_w, 5, _pdf_safe(label))

    pdf.set_fill_color(*LINE_RGB)
    pdf.rect(bar_x, y + 2, bar_w, bar_h, style="F")

    if has_data and score is not None:
        filled_w = bar_w * max(0, min(1, score / 10))
        pdf.set_fill_color(*ACCENT_RGB)
        if filled_w > 0:
            pdf.rect(bar_x, y + 2, filled_w, bar_h, style="F")
        pdf.set_xy(bar_x + bar_w + 3, y + 1)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*INK_RGB)
        pdf.cell(19, 5, f"{score:.1f}/10")
    else:
        pdf.set_xy(bar_x + bar_w + 3, y + 1)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(*FAINT_RGB)
        pdf.cell(19, 5, "No data")

    pdf.set_text_color(*INK_RGB)
    pdf.set_y(y + 8)


def _latest_coaching_plan(user_id: str) -> Optional[dict]:
    rows = supabase_client.db_select(
        "coaching_plans",
        params={
            "user_id": f"eq.{user_id}",
            "select": "*",
            "order": "created_at.desc",
            "limit": "1",
        },
    )
    return rows[0] if rows else None


def _pdf_safe(text) -> str:
    """
    fpdf2's built-in core fonts (Helvetica/Times/Courier) only support
    Latin-1 - any other character (smart quotes, em/en dashes, bullets,
    emoji, non-Latin scripts) raises FPDFUnicodeEncodingException and
    fails the ENTIRE PDF, not just that one piece of text. AI-generated
    text routinely contains exactly these characters (confirmed: an
    em-dash alone is enough to reproduce this).

    Every piece of dynamic text going into the PDF - user names,
    coaching summaries, activity titles - passes through this first.
    Common typographic characters get mapped to plain ASCII
    equivalents; anything still left outside Latin-1 gets replaced
    rather than crashing the whole report.
    """
    if not text:
        return ""
    text = str(text)
    replacements = {
        "\u2018": "'", "\u2019": "'",  # smart single quotes
        "\u201c": '"', "\u201d": '"',  # smart double quotes
        "\u2013": "-", "\u2014": "-",  # en dash, em dash
        "\u2026": "...",  # ellipsis
        "\u2022": "-",  # bullet
        "\u00a0": " ",  # non-breaking space
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _activity_title(item: dict) -> str:
    text = item.get("topic") or item.get("input_text") or item.get("transcript") or ""
    return text[:70]


def build_progress_pdf(user_id: str, full_name: str) -> bytes:
    performance = compute_performance_score(user_id)
    activity = get_recent_activity(user_id, per_source_limit=10)
    coaching = _latest_coaching_plan(user_id)

    pdf = ClashLabPDF()
    pdf.add_page()
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    _draw_header(pdf, "Progress Report", f"{full_name}  -  generated {generated}")

    overall = performance["overall_score"]
    overall_text = f"{overall:.1f}/10" if overall is not None else "Not enough data yet"
    _score_badge(pdf, "Overall Performance Score", overall_text)
    pdf.ln(4)

    _section_heading(pdf, "Component Breakdown")
    for c in performance["components"]:
        _score_bar_row(pdf, f"{c['label']} ({c['weight_pct']}%)", c["score"], c["has_data"])
    pdf.ln(2)

    if coaching:
        _section_heading(pdf, "Latest Coaching Summary")
        pdf.set_font("Helvetica", "", 10)
        summary = _pdf_safe(coaching.get("summary_feedback", "")) or "(no summary)"
        pdf.multi_cell(0, 6, summary, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(2)

    _section_heading(pdf, "Recent Activity")
    pdf.set_font("Helvetica", "", 9)
    if not activity:
        pdf.set_text_color(*FAINT_RGB)
        pdf.cell(0, 6, "No activity yet.", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_text_color(*INK_RGB)
    for item in activity[:25]:
        date = (item.get("created_at") or "")[:10]
        label = item.get("tool_label", item.get("kind", ""))
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(*ACCENT_RGB)
        pdf.write(5, _pdf_safe(f"[{label}] "))
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*INK_RGB)
        pdf.multi_cell(0, 5, _pdf_safe(f"{date} - {_activity_title(item)}"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    return bytes(pdf.output())


def build_progress_excel(user_id: str, full_name: str) -> bytes:
    performance = compute_performance_score(user_id)

    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"

    overview["A1"] = "ClashLab Progress Report"
    overview["A1"].font = Font(bold=True, size=14)
    overview["A2"] = "Name"
    overview["B2"] = full_name
    overview["A3"] = "Generated"
    overview["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    overview["A4"] = "Overall Score"
    overview["B4"] = performance["overall_score"]

    overview["A6"] = "Component"
    overview["B6"] = "Weight %"
    overview["C6"] = "Score"
    overview["D6"] = "Has Data"
    for cell in ("A6", "B6", "C6", "D6"):
        overview[cell].font = Font(bold=True)
    for i, c in enumerate(performance["components"], start=7):
        overview[f"A{i}"] = c["label"]
        overview[f"B{i}"] = c["weight_pct"]
        overview[f"C{i}"] = c["score"]
        overview[f"D{i}"] = c["has_data"]

    for sheet_name, table, columns in TOOL_SHEETS:
        rows = supabase_client.db_select(
            table,
            params={
                "user_id": f"eq.{user_id}",
                "select": "*",
                "order": "created_at.desc",
                "limit": "200",
            },
        )
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(col) for col in columns])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- Per-item reports (Segment 17) ---
#
# Rather than a bespoke PDF layout for each of the 6+ different data
# shapes (an argument analysis, a debate session with nested rounds
# and judge feedback, a coaching plan, ...), one generic recursive
# renderer handles all of them: dicts become labeled sub-sections,
# lists become bulleted (or numbered, for lists of objects) entries,
# scalars print as "Label: value". This is far more maintainable than
# hand-writing a layout per tool, and new fields any tool adds later
# automatically show up with no report code changes needed.

ITEM_SKIP_FIELDS = {"id", "user_id", "session_id", "created_at"}
ITEM_PRIORITY_ORDER = [
    "topic", "input_text", "transcript", "overall_score", "credibility_score",
]


def _humanize_key(key: str) -> str:
    words = key.replace("_", " ").split(" ")
    return " ".join(w.upper() if w.lower() == "ai" else w.capitalize() for w in words)


def _render_value(pdf: FPDF, key, value, indent: int = 0) -> None:
    prefix = "  " * indent
    label = _humanize_key(key) if key else None

    if isinstance(value, dict):
        if label:
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*ACCENT_RGB)
            pdf.multi_cell(0, 6, _pdf_safe(f"{prefix}{label}:"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_text_color(*INK_RGB)
            pdf.set_font("Helvetica", "", 9)
        for k, v in value.items():
            if k in ITEM_SKIP_FIELDS:
                continue
            _render_value(pdf, k, v, indent + 1)

    elif isinstance(value, list):
        if not value:
            return
        if label:
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*ACCENT_RGB)
            pdf.multi_cell(0, 6, _pdf_safe(f"{prefix}{label}:"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_text_color(*INK_RGB)
            pdf.set_font("Helvetica", "", 9)
        for i, v in enumerate(value):
            if isinstance(v, (dict, list)):
                pdf.set_font("Helvetica", "B", 9)
                pdf.multi_cell(
                    0, 5, _pdf_safe(f"{prefix}  Item {i + 1}:"),
                    new_x=XPos.LMARGIN, new_y=YPos.NEXT,
                )
                pdf.set_font("Helvetica", "", 9)
                _render_value(pdf, None, v, indent + 2)
            else:
                pdf.multi_cell(
                    0, 5, _pdf_safe(f"{prefix}  - {v}"), new_x=XPos.LMARGIN, new_y=YPos.NEXT
                )

    else:
        if value is None or value == "":
            return
        text = f"{prefix}{label}: {value}" if label else f"{prefix}{value}"
        pdf.multi_cell(0, 5, _pdf_safe(str(text)), new_x=XPos.LMARGIN, new_y=YPos.NEXT)


def build_item_pdf(tool_label: str, item: dict, generated_for: str) -> bytes:
    pdf = ClashLabPDF()
    pdf.add_page()
    created = (item.get("created_at") or "")[:10]
    _draw_header(pdf, f"{tool_label} Report", f"{generated_for}  -  {created}")
    pdf.set_font("Helvetica", "", 9)

    ordered_keys = [k for k in ITEM_PRIORITY_ORDER if k in item] + [
        k for k in item if k not in ITEM_PRIORITY_ORDER and k not in ITEM_SKIP_FIELDS
    ]

    for key in ordered_keys:
        if key in ITEM_SKIP_FIELDS:
            continue
        _render_value(pdf, key, item[key])
        pdf.ln(2)

    return bytes(pdf.output())


async def build_class_report_pdf(coach_name: str) -> bytes:
    """
    Segment 20: a coach/educator-level report, one page listing every
    learner on the platform. Thin wrapper around build_roster_report_pdf
    below (Segment 21) - kept as its own function since it's the
    existing entry point /reports/class/pdf already calls.
    """
    from app.services.dashboard_service import get_all_learners_overview_async

    learners = await get_all_learners_overview_async()
    return build_roster_report_pdf(coach_name, learners, "Class Report")


def build_roster_report_pdf(coach_name: str, learners: list, title: str) -> bytes:
    """
    The actual rendering logic, extracted in Segment 21 so a
    class-scoped report (a specific class's roster) and the
    platform-wide report (every learner) share the exact same layout
    instead of two near-identical copies.
    """
    pdf = ClashLabPDF()
    pdf.add_page()
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    _draw_header(pdf, title, f"Prepared for {coach_name}  -  {generated}")

    _section_heading(pdf, f"{len(learners)} Learner(s), Ranked by Overall Score")

    for i, learner in enumerate(learners, start=1):
        overall_text = (
            f"{learner['overall_score']:.1f}/10" if learner["overall_score"] is not None else "No data yet"
        )
        pdf.set_font("Helvetica", "B", 10)
        pdf.multi_cell(
            0, 6, _pdf_safe(f"{i}. {learner['full_name']}"), new_x=XPos.LMARGIN, new_y=YPos.NEXT
        )
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*FAINT_RGB)
        experience = learner.get("experience_level") or "Unspecified experience"
        counts = learner.get("data_counts", {})
        counts_text = ", ".join(f"{v} {k.replace('_', ' ')}" for k, v in counts.items() if v)
        pdf.multi_cell(
            0, 5,
            _pdf_safe(f"   Overall: {overall_text}  |  {experience}" + (f"  |  {counts_text}" if counts_text else "")),
            new_x=XPos.LMARGIN, new_y=YPos.NEXT,
        )
        pdf.set_text_color(*INK_RGB)
        pdf.ln(2)

    if not learners:
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*FAINT_RGB)
        pdf.cell(0, 6, "No learners here yet.", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_text_color(*INK_RGB)

    return bytes(pdf.output())
