import csv
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Response, status
from sqlalchemy.orm import Session

import models
from database import get_db
from routers.auth import get_current_user
from routers.coaching import build_coaching_plan

router = APIRouter(prefix="/api/v1/reports", tags=["Reports & Export System"])


def _owned_session(session_id: int, user: models.User, db: Session) -> models.DebateSession:
    session = (
        db.query(models.DebateSession)
        .filter(models.DebateSession.id == session_id, models.DebateSession.user_id == user.id)
        .first()
    )
    if not session:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Debate session not found.")
    return session


def _session_data(session_id: int, user: models.User, db: Session) -> dict:
    session = _owned_session(session_id, user, db)
    score = (
        db.query(models.PerformanceScore)
        .filter(models.PerformanceScore.session_id == session.id, models.PerformanceScore.user_id == user.id)
        .order_by(models.PerformanceScore.created_at.desc())
        .first()
    )
    metric = (
        db.query(models.PresentationMetric)
        .filter(models.PresentationMetric.session_id == session.id, models.PresentationMetric.user_id == user.id)
        .order_by(models.PresentationMetric.created_at.desc())
        .first()
    )
    analyses = (
        db.query(models.ArgumentAnalysis)
        .filter(models.ArgumentAnalysis.session_id == session.id, models.ArgumentAnalysis.user_id == user.id)
        .order_by(models.ArgumentAnalysis.created_at.asc())
        .all()
    )
    turns = (
        db.query(models.SimulationTurn)
        .filter(models.SimulationTurn.session_id == session.id, models.SimulationTurn.user_id == user.id)
        .order_by(models.SimulationTurn.turn_index.asc())
        .all()
    )
    feedback = (
        db.query(models.CoachFeedback)
        .filter(models.CoachFeedback.session_id == session.id, models.CoachFeedback.learner_id == user.id)
        .order_by(models.CoachFeedback.created_at.desc())
        .all()
    )
    fallacies = [fallacy.fallacy_type for analysis in analyses for fallacy in analysis.fallacies]
    return {
        "session": session,
        "score": score,
        "metric": metric,
        "analyses": analyses,
        "turns": turns,
        "feedback": feedback,
        "fallacies": fallacies,
    }


def _pdf_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def generate_pdf_bytes(title: str, content_lines: list[str]) -> bytes:
    # Small dependency-free PDF writer for text reports.
    stream = ["BT", "/F1 16 Tf", "50 790 Td", f"({_pdf_escape(title)}) Tj", "/F1 10 Tf"]
    for line in content_lines:
        stream.extend(["0 -18 Td", f"({_pdf_escape(line[:160])}) Tj"])
    stream.append("ET")
    content = "\n".join(stream).encode("latin-1", "replace")
    objects = [
        b"<< /Type /Catalog /Pages 3 0 R >>",
        b"<< /Type /Outlines /Count 0 >>",
        b"<< /Type /Pages /Kids [4 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 3 0 R /MediaBox [0 0 595 842] /Contents 5 0 R /Resources << /Font << /F1 6 0 R >> >> >>",
        b"<< /Length " + str(len(content)).encode() + b" >>\nstream\n" + content + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    body = [b"%PDF-1.4\n"]
    offsets = [0]
    current = len(body[0])
    for index, obj in enumerate(objects, 1):
        offsets.append(current)
        block = f"{index} 0 obj\n".encode() + obj + b"\nendobj\n"
        body.append(block)
        current += len(block)
    xref_offset = current
    body.append(b"xref\n0 7\n0000000000 65535 f \n")
    body.extend(f"{offset:010d} 00000 n \n".encode() for offset in offsets[1:])
    body.extend([
        b"trailer\n<< /Size 7 /Root 1 0 R >>\nstartxref\n",
        str(xref_offset).encode(),
        b"\n%%EOF\n",
    ])
    return b"".join(body)


def _report_lines(data: dict) -> list[str]:
    session = data["session"]
    score = data["score"]
    metric = data["metric"]
    lines = [
        "DEBATE SESSION",
        f"Topic: {session.topic}",
        f"Title: {session.title}",
        f"Format: {session.format} | Position: {session.assigned_position}",
        f"Status: {session.status}",
        "",
        "PERFORMANCE SCORE",
        f"Overall weighted score: {score.overall_weighted_score:.2f}%" if score else "Overall weighted score: Not scored",
        f"Argument quality: {score.argument_quality:.2f}%" if score else "Argument quality: Not scored",
        f"Evidence use: {score.evidence_use:.2f}%" if score else "Evidence use: Not scored",
        f"Logical consistency: {score.logical_consistency:.2f}%" if score else "Logical consistency: Not scored",
        f"Rebuttal effectiveness: {score.rebuttal_effectiveness:.2f}%" if score else "Rebuttal effectiveness: Not scored",
        f"Communication skills: {score.communication_skills:.2f}%" if score else "Communication skills: Not scored",
        "",
        "PRESENTATION METRICS",
        f"Speaking pace: {metric.speech_pace_wpm:.2f} WPM" if metric else "Speaking pace: Not analyzed",
        f"Filler words: {metric.filler_words_count}" if metric else "Filler words: Not analyzed",
        f"Clarity: {metric.clarity_score:.2f}%" if metric else "Clarity: Not analyzed",
        f"Confidence: {metric.confidence_score:.2f}%" if metric else "Confidence: Not analyzed",
        f"Engagement: {metric.engagement_score:.2f}%" if metric else "Engagement: Not analyzed",
        "",
        "FALLACIES DETECTED",
        ", ".join(data["fallacies"]) if data["fallacies"] else "None recorded",
        "",
        "SIMULATION TURNS",
    ]
    for turn in data["turns"]:
        lines.extend([
            f"Turn {turn.turn_index} // {turn.opponent_persona}",
            f"User argument: {turn.user_argument}",
            f"Opponent rebuttal: {turn.opponent_rebuttal}",
            f"Rebuttal strength: {turn.rebuttal_strength_percent:.1f}% | Coaching: {turn.coaching_tip}",
        ])
    if data["feedback"]:
        lines.extend(["", "COACH FEEDBACK"])
        for item in data["feedback"]:
            rating = f" ({item.rating:.1f}/100)" if item.rating is not None else ""
            lines.append(f"{item.content}{rating}")
    lines.append(f"Generated at: {datetime.now(timezone.utc).replace(tzinfo=None).isoformat(timespec='seconds')}Z")
    return lines


@router.get("/export/pdf/{session_id}")
def export_pdf_report(
    session_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    data = _session_data(session_id, current_user, db)
    return Response(
        content=generate_pdf_bytes(f"LOGOS.AI ASSESSMENT REPORT // SESSION {session_id}", _report_lines(data)),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=logos_ai_session_{session_id}_assessment.pdf"},
    )


@router.get("/export/excel/{session_id}")
def export_excel_report(
    session_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    data = _session_data(session_id, current_user, db)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise HTTPException(status_code=503, detail="XLSX export dependency is not installed.") from exc

    session, score, metric = data["session"], data["score"], data["metric"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Session Report"
    sheet.append(["LOGOS.AI SESSION REPORT"])
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append(["Session ID", session.id])
    sheet.append(["Topic", session.topic])
    sheet.append(["Title", session.title])
    sheet.append(["Format", session.format])
    sheet.append(["Position", session.assigned_position])
    sheet.append(["Status", session.status])
    sheet.append([])
    sheet.append(["Metric", "Value", "Weight"])
    if score:
        for label, value, weight in [
            ("Argument quality", score.argument_quality, "30%"),
            ("Evidence use", score.evidence_use, "20%"),
            ("Logical consistency", score.logical_consistency, "20%"),
            ("Rebuttal effectiveness", score.rebuttal_effectiveness, "15%"),
            ("Communication skills", score.communication_skills, "15%"),
            ("Overall weighted score", score.overall_weighted_score, "100%"),
        ]:
            sheet.append([label, value, weight])
    if metric:
        sheet.append(["Speaking pace (WPM)", metric.speech_pace_wpm, "N/A"])
        sheet.append(["Filler words", metric.filler_words_count, "N/A"])
        sheet.append(["Clarity", metric.clarity_score, "N/A"])
        sheet.append(["Confidence", metric.confidence_score, "N/A"])
        sheet.append(["Engagement", metric.engagement_score, "N/A"])
    turns_sheet = workbook.create_sheet("Simulation Turns")
    turns_sheet.append(["Turn", "Persona", "User argument", "Opponent rebuttal", "Strength", "Coaching tip"])
    for turn in data["turns"]:
        turns_sheet.append([turn.turn_index, turn.opponent_persona, turn.user_argument, turn.opponent_rebuttal, turn.rebuttal_strength_percent, turn.coaching_tip])
    feedback_sheet = workbook.create_sheet("Coach Feedback")
    feedback_sheet.append(["Coach ID", "Rating", "Feedback", "Created at"])
    for item in data["feedback"]:
        feedback_sheet.append([item.coach_id, item.rating, item.content, item.created_at.isoformat() if item.created_at else ""])
    for current_sheet in workbook.worksheets:
        for column in current_sheet.columns:
            current_sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
    output = io.BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=logos_ai_session_{session_id}_report.xlsx"},
    )


@router.get("/export/coaching/pdf/{user_id}")
def export_coaching_pdf_report(
    user_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You can only export your own coaching plan.")
    plan = build_coaching_plan(current_user.id, db)
    lines = [
        "COACHING PROFILE & PROGRESS SUMMARY",
        f"User: {current_user.full_name}",
        f"Status: {plan['progress_status']}",
        "",
        "SKILL GAP ANALYSIS",
        plan["skill_gap_summary"],
        "",
        "TARGETED RECOMMENDATIONS",
        *[f"- {recommendation}" for recommendation in plan["targeted_recommendations"]],
        "",
        "LEARNING PATH",
        *[f"- {step}" for step in plan["learning_path_steps"]],
    ]
    return Response(
        content=generate_pdf_bytes("LOGOS.AI COACHING PROGRESS REPORT", lines),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=logos_ai_user_{user_id}_coaching_plan.pdf"},
    )


@router.get("/export/summary/{session_id}")
def get_session_summary_report(
    session_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    data = _session_data(session_id, current_user, db)
    score, metric = data["score"], data["metric"]
    return {
        "platform": "LOGOS.AI",
        "session_id": session_id,
        "title": data["session"].title,
        "topic": data["session"].topic,
        "status": data["session"].status,
        "weighted_performance_score": score.overall_weighted_score if score else None,
        "fallacies_detected": sorted(set(data["fallacies"])),
        "speech_pace_wpm": metric.speech_pace_wpm if metric else None,
        "filler_words_count": metric.filler_words_count if metric else None,
        "certificate_id": f"CERT-LOGOS-{session_id}" if score else None,
    }
